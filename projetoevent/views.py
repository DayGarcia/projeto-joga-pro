import json
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.views.generic import View
from django.core import serializers

import openpyxl
from projetoevent.models import Event, Ticket, TicketLog
from projetoevent.helpers.ticket_query import format_event_to_json, get_ticket_data
from projetoevent.helpers.event_query import get_running_event_id


class Home(View):
    def get(self, request):

        context = get_ticket_data(get_running_event_id(), request.GET.get(
            'gate_id', None), request.GET.get('user', None), request.GET.get('ticket', None))

        return render(request, 'projetoevent/home.html', context)

    def post(self, request):
        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook(request.FILES['file'])

        # Define variable to read the active sheet:
        worksheet = wookbook.active

        e = Event(
            is_running=0
        )
        e.save()

        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row):
            if i == 0:
                continue
            for col in worksheet.iter_cols(1, worksheet.max_column):
                t = Ticket(
                    code=worksheet.cell(row=i+1, column=2).value,
                    rfid=worksheet.cell(row=i+1, column=3).value,
                    status=worksheet.cell(row=i+1, column=4).value,
                    user_id=worksheet.cell(row=i+1, column=5).value,
                    type=worksheet.cell(row=i+1, column=6).value,
                    gate=worksheet.cell(row=i+1, column=7).value,
                    sector=worksheet.cell(row=i+1, column=8).value,
                    block=worksheet.cell(row=i+1, column=9).value,
                    row=worksheet.cell(row=i+1, column=10).value,
                    seat=worksheet.cell(row=i+1, column=11).value,
                    extra=worksheet.cell(row=i+1, column=12).value,
                    event=e
                )
                t.save()

        return render(request, 'projetoevent/home.html', {'msg': 'Jogo importado com sucesso!'})
        # return redirect('/home', {'msg': 'Jogo importado com sucesso!'})


class RunEvent(View):

    def post(self, request):
        # set all current event to not running
        Event.objects.filter(is_running=1).update(is_running=0)

        # set the new event to running
        e = Event.objects.get(id=request.POST['event_id'])
        e.is_running = 1
        e.save()

        # filter = '?event_id=' + request.POST['event_id']
        filter = '?filter=1'
        if(request.POST['gate_id'] != '0'):
            filter += '&gate_id=' + request.POST['gate_id']

        if(request.POST['user'] != ''):
            filter += '&user=' + request.POST['user']

        if(request.POST['ticket'] != ''):
            filter += '&ticket=' + request.POST['ticket']

        return redirect('/home' + filter, {'msg': 'Jogo iniciado com sucesso!'})


class Charts(View):
    def get(self, request):
        if 'event_id' in request.GET:
            context = get_ticket_data(request.GET['event_id'])
        else:
            context = get_ticket_data(get_running_event_id())

        return render(request, 'projetoevent/charts.html', context)


class RefreshTickets():
    def get(request):
        if is_ajax(request) and request.method == "GET":
            data = get_ticket_data(get_running_event_id())
            # json = serializers.serialize('json', data)
            jsona = format_event_to_json(data)

            return JsonResponse(jsona, status=200)

        return JsonResponse({}, status=400)


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

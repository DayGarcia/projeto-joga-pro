from projetoevent.models import Event, Ticket
import openpyxl

def save_event_task(workshives, event):

    try:
        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook(workshives)

        # Define variable to read the active sheet:
        worksheet = wookbook.active
        
        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row):
            if i == 0:
                continue
            for col in worksheet.iter_cols(1, worksheet.max_column):
                t = Ticket(
                    code=worksheet.cell(row=i + 1, column=2).value,
                    rfid=worksheet.cell(row=i + 1, column=3).value,
                    status=worksheet.cell(row=i + 1, column=4).value,
                    user_id=worksheet.cell(row=i + 1, column=5).value,
                    type=worksheet.cell(row=i + 1, column=6).value,
                    gate=worksheet.cell(row=i + 1, column=7).value,
                    sector=worksheet.cell(row=i + 1, column=8).value,
                    block=worksheet.cell(row=i + 1, column=9).value,
                    row=worksheet.cell(row=i + 1, column=10).value,
                    seat=worksheet.cell(row=i + 1, column=11).value,
                    extra=worksheet.cell(row=i + 1, column=12).value,
                    event=event
                )
            t.save()

        Event.objects.filter(is_uploading=1).update(is_uploading=0)

        return True
    except Exception as e:
        Event.objects.filter(is_uploading=1).update(is_uploading=0)
        return False
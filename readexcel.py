import openpyxl
import os
from projetoevent.models import Event, Ticket


print(os.path.join('teste.xlsx'))
# Define variable to load the wookbook
wookbook = openpyxl.load_workbook(os.path.join('teste.xlsx'))

# Define variable to read the active sheet:
worksheet = wookbook.active

e = Event()
e.save()

# Iterate the loop to read the cell values
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        t = Ticket(
            code=worksheet.cell(row=i+1, column=1).value,
            rfid=worksheet.cell(row=i+1, column=2).value,
            status=worksheet.cell(row=i+1, column=3).value,
            user_id=worksheet.cell(row=i+1, column=4).value,
            type=worksheet.cell(row=i+1, column=5).value,
            gate=worksheet.cell(row=i+1, column=6).value,
            sector=worksheet.cell(row=i+1, column=7).value,
            block=worksheet.cell(row=i+1, column=8).value,
            row=worksheet.cell(row=i+1, column=9).value,
            seat=worksheet.cell(row=i+1, column=10).value,
            extra=worksheet.cell(row=i+1, column=11).value,
            event=e
        )
